# The COPYRIGHT file at the top level of this repository contains
# the full copyright notices and license terms.
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook


def _date(value):
    v = str(value).strip()
    return datetime.strptime(v, '%Y%m%d').date()


def _string(value):
    return value.strip()


def _amount(value):
    if not value:
        return Decimal('0.0')
    return Decimal(str(value).replace(',', '.'))


def _code_to_string(value):
    try:
        return str(int(value))
    except ValueError:
        return str(value)


COL = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6}

MOVE = {
    'date': ('A', _date),
    'concept': ('B', _string),
    'debit': ('D', _amount),
    'credit': ('E', _amount),
    'code': ('G', _code_to_string),
    }


class Credicoop(object):

    def __init__(self, name, encoding='windows-1252'):
        self.statements = []
        self._parse(name)

    def _parse(self, infile):
        statement = Statement()
        self.statements.append(statement)

        filedata = BytesIO(infile)
        wb = load_workbook(filedata)
        sheet = wb.active
        for row in sheet.iter_rows():
            # Start reading at the second row
            if row[0].row == 1:
                continue
            move = Move()
            self._parse_move(row, move, MOVE)
            # 'Date from', in first row
            if row[0].row == 2:
                statement.date_from = move.date
            statement.date_to = move.date
            move.op_number = row[0].row
            statement.moves.append(move)

    def _parse_move(self, row, move, desc):
        for name, (col, parser) in desc.items():
            value = parser(row[COL[col]].value)
            setattr(move, name, value)


class Statement(object):
    __slots__ = ['date_from', 'date_to', 'moves']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.moves = []


class Move(object):
    __slots__ = list(MOVE.keys()) + ['op_number']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
